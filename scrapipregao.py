import time
import random
import openpyxl
import PySimpleGUI as sg
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook, load_workbook, writer
from openpyxl.styles import Alignment, NamedStyle, PatternFill
from openpyxl.worksheet.filters import FilterColumn, CustomFilter, CustomFilters, DateGroupItem, Filters


navegador = webdriver.Chrome()
############# FUNÇÕES #####################
### formatação da tabela
def centralizando(w):
    alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    for row in w.iter_rows():
        for cell in row:
            cell.alignment = alignment

def tamanhoColunaComum(a):
    for col in a.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        a.column_dimensions[column].width = adjusted_width


##########################         criando esperas para o código        ############################

def wait60(navegador,urlAtual):
    try:
        wait60 = WebDriverWait(navegador, 60) 
        wait60.until(EC.url_changes(urlAtual))
        time.sleep(3)
    except:
        navegador.execute_script("var e = alert('o tempo de carregamento da pagina passou do limite, reinicie o app.'), '');document.body.setAttribute('tempoEspirador', g)")


def randomWait(navegador):
    try:
        aleatorio = random.randint(5, 10)
        print(f'inicio da espera aleatoria de {aleatorio} segundos')
        time.sleep(aleatorio)

    except:
        navegador.execute_script("var e = alert('o tempo de carregamento da pagina passou do limite, reinicie o app.'), '');document.body.setAttribute('tempoEspirador', g)")

#########################        função para abrir a aba empresas e retirar as informações de cada empresa resumida       ############################
def informaçoesEmpresas(qntEmpresas):
    informaçoesEmpresas = []
    for j in range (1, qntEmpresas+1):
        cnpjEmpresaCompleto = navegador.find_element(By.XPATH, '/html/body/app-root/div/div/div/app-cabecalho-selecao-fornecedores-governo/div[2]/app-selecao-fornecedores-governo/div/p-tabview/div/div[2]/p-tabpanel[2]/div/app-selecao-fornecedores-governo-participantes/div[2]/p-dataview/div/div/div['+ str(j) + ']/div[1]/div/div[1]').text
        nomeEmpresa = navegador.find_element(By.XPATH, '/html/body/app-root/div/div/div/app-cabecalho-selecao-fornecedores-governo/div[2]/app-selecao-fornecedores-governo/div/p-tabview/div/div[2]/p-tabpanel[2]/div/app-selecao-fornecedores-governo-participantes/div[2]/p-dataview/div/div/div[' + str(j) + ']/div[2]/div/span').text        
        if len(cnpjEmpresaCompleto.split('\n')) == 1:
            cnpjEmpresa = cnpjEmpresaCompleto
            informacoesEmpresa = {'CNPJ': cnpjEmpresa, 'Nome': nomeEmpresa, 'ME/EPP': 'Não'}
        else:
            cnpjEmpresa = cnpjEmpresaCompleto.split('\n')[0]
            informacoesEmpresa = {'CNPJ': cnpjEmpresa, 'Nome': nomeEmpresa, 'ME/EPP': 'Sim'}
        informaçoesEmpresas.append(informacoesEmpresa)  
    return informaçoesEmpresas

##########################        função para abrir os itens e retirar as informações necessárias        ############################
def informaçoesItens():
    #descrição do item:
    descricaoResumidaItem = navegador.find_element(By.XPATH, '/html/body/app-root/div/div/div/app-cabecalho-selecao-fornecedores-governo/div[2]/app-selecao-fornecedores-governo-item/div/div/app-cabecalho-item/div/div[1]/div/app-identificacao-e-fase-item/div[1]').text
    #valor estimado:
    valorEstimado = navegador.find_element(By.XPATH, '/html/body/app-root/div/div/div/app-cabecalho-selecao-fornecedores-governo/div[2]/app-selecao-fornecedores-governo-item/div/div/app-cabecalho-item/div/div[2]/div/div/div/div[2]/div[2]').text.split(' ')[1]
    #qnt solicitada:
    qntSolicitada = navegador.find_element(By.XPATH, '/html/body/app-root/div/div/div/app-cabecalho-selecao-fornecedores-governo/div[2]/app-selecao-fornecedores-governo-item/div/div/app-cabecalho-item/div/div[2]/div/div/div/div[2]/div[1]').text
    return descricaoResumidaItem, valorEstimado, qntSolicitada

def proposta(i):
    cnpjMEPP = navegador.find_element(By.XPATH, '/html/body/app-root/div/div/div/app-cabecalho-selecao-fornecedores-governo/div[2]/app-selecao-fornecedores-governo-item/div/div/app-selecao-fornecedores-governo-propostas-item/div/div/div/p-dataview/div/div/div[' + str(i) + ']/app-dados-proposta-item-em-selecao-fornecedores/div/div[1]/div/app-identificacao-e-situacao-participante-no-item/div/div[1]').text.split('\n')
    #nome empresa:
    nomeEmpresa = navegador.find_element(By.XPATH, '/html/body/app-root/div/div/div/app-cabecalho-selecao-fornecedores-governo/div[2]/app-selecao-fornecedores-governo-item/div/div/app-selecao-fornecedores-governo-propostas-item/div/div/div/p-dataview/div/div/div[' + str(i) + ']/app-dados-proposta-item-em-selecao-fornecedores/div/div[1]/div/app-identificacao-e-situacao-participante-no-item/div/div[2]/div[1]/span').text
    #valor ofertado pela empresa:
    valorOfertado = navegador.find_element(By.XPATH, '/html/body/app-root/div/div/div/app-cabecalho-selecao-fornecedores-governo/div[2]/app-selecao-fornecedores-governo-item/div/div/app-selecao-fornecedores-governo-propostas-item/div/div/div/p-dataview/div/div/div[' + str(i) + ']/app-dados-proposta-item-em-selecao-fornecedores/div/div[2]/div/div/div[2]/div[1]/span/span').text.split(' ')[1]
    return cnpjMEPP, nomeEmpresa, valorOfertado


def formatarValor(valor):
    valor = valor.replace('.', '').replace(',', '.')
    valor = valor.replace('.', ',')
    return valor

options = Options()
options.page_load_strategy = 'normal'


##########################         Login no site         ############################
# layout = [
#     [sg.Text('Para a criação da planilha, preencha os campos abaixo conforme os exemplos;')],
#     [sg.Text('Após o ok irá abrir o browser e deverá ser efetuado o login no site do compras;')],
#     [sg.Text('Durante a execução do processo, não click em nada na janela aberta pelo programa, o restante pode ser usado normalmente;')],
#     [sg.Text('Trata-se de um programa de teste em aperfeiçoamento e elaboração pelo Ten Fazzolo, confira os valores da planilha com o sistema antes de qualquer atitude;')],
#     [sg.Text('Esta versão está funcionando apenas para pregões sem grupos, uma nova versao para grupos será desenvolvida futuramente;')],
#     [sg.Text('Acorreta execução do programa depende da velocidade do computador e internet, caso ocorra algum erro, tente em outro computador, caso persista, enviar um email para admfazzolo@gmail.com com a pane apresentada.')],
#     [sg.Text('UASG:'), sg.InputText(key='uasg')],
#     [sg.Text('Número do pregão (ex: XX/202X):'), sg.InputText(key='numero')],
#     [sg.Text('Quantidade de Empresas:'), sg.InputText(key='qntEmpresas')],
#     [sg.Button('Sair'), sg.Button('Enviar')]
# ]

# window = sg.Window('Dados do Pregão', layout)
# while True:
#     event, values = window.read()

#     if event in (sg.WIN_CLOSED, 'Sair'):
#         break
#     elif event == 'Enviar':
#         uasg = values['uasg']
#         qntEmpresas = int(values['qntEmpresas'])
#         numero = values['numero']

#         confirmationLayout = [
#             [sg.Text(f'UASG: {uasg}')],
#             [sg.Text(f'Número do pregão: {numero}')],
#             [sg.Text(f'Quantidade de Empresas: {qntEmpresas}')],
#             [sg.Button('Alterar'), sg.Button('Confirmar')],
#         ]

#         confirmationWindow = sg.Window('Confirmação', confirmationLayout)

#         while True:
#             event, _ = confirmationWindow.read()
#             if event == 'Alterar':
#                 confirmationWindow.close()
#                 break
#             elif event == 'Confirmar':
#                 confirmationWindow.close()
#                 window.close()
#                 break

qntEmpresas = 15
uasg = str(120071)
numero = "90027/2024"
pregao = "Pregão Eletrônico " + uasg + " - " + numero


time.sleep(10)
navegador.get('https://cnetmobile.estaleiro.serpro.gov.br/comprasnet-area-trabalho-web/seguro/governo/area-trabalho')
urlAtual = navegador.current_url
time.sleep(10)

##########################         espera para o login do usuario de 5min       ############################
try:
    waitLogin = WebDriverWait(navegador, 900).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="buscaCompra"]'))
    )
except:
    navegador.execute_script("var e = alert('Sistema encerrando pois usuario não efetuou login, tente novamente e efetue o login.', '');document.body.setAttribute('loginscrapynãofeito', f)")
    time.sleep(10)

##########################        CRIANDO A TABELA DE APOIO        ############################

wb = openpyxl.Workbook()
####       CRIANDO A ABA COM TODAS EMPRESAS 

wb.create_sheet('Empresas',0)
wbEmpresas = wb['Empresas']
#cabeçalho
wbEmpresas.append(['CNPJ', 'ME/EPP', 'NOME EMPRESA'])

####       CRIANDO A ABA DOS ITENS COM AS INFORMAÇÕES DAS 3 PRIMEIRAS EMPRESAS      

wb.create_sheet('Três primeiras empresas',1)
wbItens = wb['Três primeiras empresas']
#cabeçalho
wbItens.append(['Item/Descrição Resumida', 'Valor Estimado','Qnt Solicitada', 'Empresa', 'Valor ofertado pela empresa'])

####      CRIANDO A ABA DE ANÁLISE DA EMPRESA 

wb.create_sheet('Análise da empresa',2)
wbAnalise = wb['Análise da empresa']
#cabeçalho
wbAnalise.append(['Descrição Resumida',	'Empresa', 'Qnt Solicitada', 'Valor Estimado', 'Valor ofertado pela empresa', 'Negociou Valor?', 'Especificação Técnica', 'Validade da Proposta', 'SICAF', 'Sanção / Ocorrência', 'CEIS', 'CNJ', 'TCU', 'Empresário Individual:Inscrição no Registro Público', 'MEI:Certificado da Condição de Microempreendedor Individual Verificar autenticidade', 'Sociedade Empresária ou Empresa Individual de Responsabilidade Limitada: Ato Constitutivo, Estatuto ou Contrato Social', 'Ato Constitutivo', 'Inscrição CNPJ', 'Regularidade Fiscal Fazenda Nacional', 'FGTS', 'CNDT', 'Inscrição Contribuintes Estadual -Excluído para ME/EPP', 'Regularidade Fazenda Estadual - Excluído para ME/EPP', 'Certidão Negativa de Falência', 'Balanço Patrimonial - Excluído para ME/EPP', 'Boa Situação Financeira', 'Habilitação Técnica'])

numero1 = numero.replace('/', '_')
arquivoExcel = 'Planilha Apoio Pregao '+ str(numero1)+'.xlsx'
wb.save(arquivoExcel)


##########################         abrindo o pregao 14.133        ############################
wait60(navegador,urlAtual)
urlAtual = navegador.current_url
navegador.find_element(By.XPATH, '//*[@id="buscaCompra"]').send_keys(numero)
navegador.find_element(By.XPATH, '/html/body/app-root/div/app-area-governo/div/div[2]/div[2]/div/div[2]/button').click()
wait60(navegador,urlAtual)
for i in range(10):
    urlXPATH = str('/html/body/app-root/div/app-pesquisa-rapida/div/div[5]/p-dataview/div/div['+ str(i) +']/div/div/div[2]/span')
    pregoesPesquisa = navegador.find_elements(By.XPATH, urlXPATH )
    for e in pregoesPesquisa:
        a = 1
        if e.text == str(pregao):
            bottonXpath = str('/html/body/app-root/div/app-pesquisa-rapida/div/div[5]/p-dataview/div/div[2]/div['+ str(a) +']/div/div[5]/i')
            navegador.find_element(By.XPATH, bottonXpath).click()
            time.sleep(3)
            navegador.find_element(By.XPATH, '/html/body/app-root/div/app-pesquisa-rapida/div/p-dialog/div/div/div[2]/app-item-trabalho-detalhe/div[2]/div[2]/div/div[2]/app-redirect-sistemas/span/span').click()
            break
        else:
            a =+ 1 
            continue
        break



time.sleep(5)                                 
##########################         alterando janela        ############################
handles = navegador.window_handles
navegador.switch_to.window(handles[1])




######## ABRINDO ABA EMPRESAS ######################
time.sleep(5)
urlAtual = navegador.current_url
navegador.find_element(By.XPATH, '/html/body/app-root/div/div/div/app-cabecalho-selecao-fornecedores-governo/div[2]/app-selecao-fornecedores-governo/div/p-tabview/div/div[1]/div/ul/li[2]/a').click()

#####       retirando as informações das empresas  

time.sleep(5)
todasEmpresas = informaçoesEmpresas(qntEmpresas)
wb = load_workbook(arquivoExcel)
wbEmpresas = wb['Empresas']
#colocando itens na planilha aba 1
linha = 2  
for a in todasEmpresas:
    cnpj = a['CNPJ']
    meEpp = a['ME/EPP']
    nome = a['Nome']
    wbEmpresas.cell(row=linha, column=1, value=cnpj)
    wbEmpresas.cell(row=linha, column=2, value=meEpp)
    wbEmpresas.cell(row=linha, column=3, value=nome)
    linha += 1
centralizando(wbEmpresas)
wbEmpresas.auto_filter.ref = wbEmpresas.dimensions
tamanhoColunaComum(wbEmpresas)
wb.save(arquivoExcel)

##########################         iterando sobre os itens do pregão        ############################
#indo para itens
time.sleep(5)
navegador.find_element(By.XPATH, '/html/body/app-root/div/div/div/app-cabecalho-selecao-fornecedores-governo/div[2]/app-selecao-fornecedores-governo/div/p-tabview/div/div[1]/div/ul/li[1]/a').click()
time.sleep(5)


#abrindo item 1
time.sleep(10)
navegador.find_element(By.XPATH, '/html/body/app-root/div/div/div/app-cabecalho-selecao-fornecedores-governo/div[2]/app-selecao-fornecedores-governo/div/p-tabview/div/div[2]/p-tabpanel[1]/div/app-selecao-fornecedores-governo-itens/div[2]/p-dataview/div/div/div[1]/app-card-item/div/div[3]/div[2]/app-botao-icone/span/button').click()
time.sleep(10)


##########################        iterando sobre todos os itens        ############################
while True:
    try:
        descricaoResumidaItem, valorEstimado, qntSolicitada = informaçoesItens()
        valorEstimado = formatarValor(valorEstimado)
        cnpjMEPP, nomeEmpresa, valorOfertado = proposta(1)
        valorOfertado = formatarValor(valorOfertado)
        wb = load_workbook(arquivoExcel)
        wbAnalise = wb['Análise da empresa']
        wbAnalise.append([descricaoResumidaItem,nomeEmpresa, qntSolicitada, valorEstimado, valorOfertado])
        wbItens = wb['Três primeiras empresas']
        wbItens.append([descricaoResumidaItem, valorEstimado, qntSolicitada, cnpjMEPP[0], nomeEmpresa, valorOfertado])
        wb.save(arquivoExcel) 
        try:
            for i in range (2, 4):
                cnpjMEPP, nomeEmpresa, valorOfertado = proposta(i)
                valorOfertado = formatarValor(valorOfertado)
                wbItens.append(['', '', '', cnpjMEPP[0], nomeEmpresa, valorOfertado])
                wb.save(arquivoExcel)    
        except:
            pass

    except:
        print('deserto')
        wbAnalise = wb['Análise da empresa']
        wbAnalise.append([descricaoResumidaItem, valorEstimado, qntSolicitada, 'DESERTO'])
        wbItens = wb['Três primeiras empresas']
        wbItens.append([descricaoResumidaItem, valorEstimado, qntSolicitada, 'DESERTO'])
        wb.save(arquivoExcel)  
        pass

    proximoItem = navegador.find_element(By.XPATH, '/html/body/app-root/div/div/div/app-cabecalho-selecao-fornecedores-governo/div[2]/app-selecao-fornecedores-governo-item/div/div/app-cabecalho-item/div/div[3]/app-botao-icone[4]/span/button')
    if proximoItem.get_attribute("disabled"):
        print('Planilha concluida')
        wb.save(arquivoExcel)  
        break
    else:
        proximoItem.click()
        randomWait(navegador)


##########################        FIM DO SCRAPY        ############################

##########################       Acertando estetica da planilha       ############################
#aba 2

centralizando(wbItens)
wbItens.auto_filter.ref = wbItens.dimensions
tamanhoColunaComum(wbItens)

#caba 3
centralizando(wbAnalise)
wbAnalise.auto_filter.ref = wbAnalise.dimensions
tamanhoColunaComum(wbAnalise)
wbAnalise.row_dimensions[1].height = 98
colunasEspecificas = ['N', 'O', 'P', 'S', 'V', 'W', 'X', 'Y','Z']

for col in colunasEspecificas:
    wbAnalise.column_dimensions[col].width = 18
    alinhamento = Alignment(horizontal='center', vertical='center',wrapText=True)
    for celula in wbAnalise[col]:
        celula.alignment = alinhamento

fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
for row in wbAnalise.iter_rows(min_row=2, min_col=4, max_col=5):
    valorColuna_D = row[0].value
    valorColuna_E = row[1].value
    if valorColuna_D < valorColuna_E:
        row[1].fill = fill


##########################        salvando planilha     ############################

wb.save(arquivoExcel)
#navegador.execute_script("var f = alert('Tabela criada com sucesso', '');document.body.setAttribute('tabelaCriada', h)")
#time.sleep(5)
#sg.popup('Planilha criada com sucesso')